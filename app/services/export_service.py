"""
Export Service — generates CSV bytes for admin cohort export.
"""

import io
import csv
from typing import List, Dict


def generate_coaching_leads_xlsx(leads: List[Dict]) -> bytes:
    """Generate an .xlsx export of coaching leads for the admin panel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Coaching Leads"

    headers = ["Captured At", "Name", "Email", "Organization", "Designation", "Country", "Phone", "Message", "Source", "Actioned"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1D3557")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for lead in leads:
        ws.append([
            str(lead.get("created_at", "") or ""),
            lead.get("name", "") or "",
            lead.get("email", "") or "",
            lead.get("organization", "") or "",
            lead.get("designation", "") or "",
            lead.get("country", "") or "",
            lead.get("phone", "") or "",
            lead.get("message", "") or "",
            lead.get("source", "") or "",
            "Yes" if lead.get("actioned") else "No",
        ])

    widths = [22, 22, 28, 24, 20, 16, 16, 50, 16, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_cohort_csv(respondents: List[Dict]) -> bytes:
    """
    Generate CSV export for a cohort.

    Each row: Name, Email, Completed, Profile (Is/Should/Want),
    P/A/E/I scores for each dimension.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Name", "Email", "Completed At",
        "Profile (Is)", "Profile (Should)", "Profile (Want)",
        "P (Is)", "A (Is)", "E (Is)", "I (Is)",
        "P (Should)", "A (Should)", "E (Should)", "I (Should)",
        "P (Want)", "A (Want)", "E (Want)", "I (Want)",
        "Dominant Style",
    ])

    for r in respondents:
        result = r.get("result") or {}
        scaled = result.get("scaled_scores", {})
        profile = result.get("profile", {})
        interp = result.get("interpretation", {})

        def sc(dim, role):
            return scaled.get(dim, {}).get(role, "")

        writer.writerow([
            r.get("name", ""),
            r.get("email", ""),
            r.get("completed_at", ""),
            profile.get("is", ""),
            profile.get("should", ""),
            profile.get("want", ""),
            sc("is", "P"), sc("is", "A"), sc("is", "E"), sc("is", "I"),
            sc("should", "P"), sc("should", "A"), sc("should", "E"), sc("should", "I"),
            sc("want", "P"), sc("want", "A"), sc("want", "E"), sc("want", "I"),
            ", ".join(interp.get("dominant_roles", [])),
        ])

    return output.getvalue().encode("utf-8")
